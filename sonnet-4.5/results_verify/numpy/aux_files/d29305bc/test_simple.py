import pandas as pd
import tempfile
import os
import sys

print(f"sys.float_info.max = {sys.float_info.max}")
print(f"Test value = 1.7976931348623155e+308")
print(f"Test value == sys.float_info.max? {1.7976931348623155e+308 == sys.float_info.max}")

df = pd.DataFrame([[1.7976931348623155e+308]])

with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
    tmp_path = tmp.name

try:
    print("\n1. Writing DataFrame to Excel...")
    df.to_excel(tmp_path, index=False)
    print("   Success!")

    print("\n2. Reading Excel file back...")
    result = pd.read_excel(tmp_path)
    print(f"   Success! Result: {result}")
except OverflowError as e:
    print(f"   OverflowError: {e}")
except Exception as e:
    print(f"   Other error: {type(e).__name__}: {e}")
finally:
    if os.path.exists(tmp_path):
        os.unlink(tmp_path)

# Let's also check what happens when we write and read using openpyxl directly
print("\n3. Checking what value is stored in Excel file...")
from openpyxl import load_workbook

df = pd.DataFrame([[1.7976931348623155e+308]])
with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
    tmp_path = tmp.name

try:
    df.to_excel(tmp_path, index=False)

    wb = load_workbook(tmp_path)
    ws = wb.active
    cell_value = ws.cell(row=2, column=1).value  # Row 2 because row 1 is header
    print(f"   Cell value from openpyxl: {cell_value}")
    print(f"   Type: {type(cell_value)}")
    print(f"   Is infinity? {cell_value == float('inf')}")

    # Try to convert to int like the bug does
    try:
        int_val = int(cell_value)
        print(f"   Converted to int: {int_val}")
    except OverflowError as e:
        print(f"   int(cell_value) raises OverflowError: {e}")

finally:
    if os.path.exists(tmp_path):
        os.unlink(tmp_path)