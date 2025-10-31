import pandas as pd
import tempfile
import os
import openpyxl

# Create a DataFrame with empty strings
df = pd.DataFrame([['', '']], columns=['A', 'B'])
print("Original DataFrame:")
print(df)
print(f"Values: {df.values.tolist()}")

with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
    filepath = f.name

# Write to Excel
df.to_excel(filepath, index=False)

# Check what's actually in the Excel file using openpyxl
wb = openpyxl.load_workbook(filepath)
ws = wb.active

print("\nExcel file contents (using openpyxl):")
print(f"Max row: {ws.max_row}")
print(f"Max column: {ws.max_column}")

for row in ws.iter_rows():
    row_data = [cell.value for cell in row]
    print(f"Row: {row_data}")

print("\nChecking specific cells:")
print(f"A1 (header): {ws['A1'].value}")
print(f"B1 (header): {ws['B1'].value}")
print(f"A2 (data): {ws['A2'].value}")
print(f"B2 (data): {ws['B2'].value}")

# Now test with na_rep parameter
print("\n" + "="*50)
print("Testing with na_rep='EMPTY':")

df.to_excel(filepath, index=False, na_rep='EMPTY')

wb2 = openpyxl.load_workbook(filepath)
ws2 = wb2.active

print("\nExcel file contents with na_rep='EMPTY':")
for row in ws2.iter_rows():
    row_data = [cell.value for cell in row]
    print(f"Row: {row_data}")

# Read back
df_read = pd.read_excel(filepath)
print("\nRead back with default settings:")
print(df_read)
print(f"Shape: {df_read.shape}")

os.unlink(filepath)